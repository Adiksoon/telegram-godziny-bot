from __future__ import annotations

import os
import re
import sqlite3
import tempfile
import logging
import threading
from http.server import BaseHTTPRequestHandler, HTTPServer
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Iterable, Iterator, Any
from zoneinfo import ZoneInfo

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import Application, CommandHandler, ContextTypes, ConversationHandler, MessageHandler, CallbackQueryHandler, filters


ROOT = Path(__file__).resolve().parent
DB_PATH = ROOT / "work_hours.sqlite3"

load_dotenv(ROOT / ".env")

BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "").strip()
HOURLY_RATE = float(os.getenv("HOURLY_RATE", "31.5").replace(",", "."))
TZ = ZoneInfo(os.getenv("TIMEZONE", "Europe/Warsaw"))
ASK_DETAILS = 0

# Stany konwersacji podsumowania dnia
END_MAIN_TASK = 1
END_ENERGY = 2
END_SENS = 3
END_FRUSTRATION = 4
END_WORK_MODE = 5

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[
        logging.FileHandler(ROOT / "bot_errors.log"),
        logging.StreamHandler()
    ]
)


DATABASE_URL = os.getenv("DATABASE_URL", "").strip()
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

IS_POSTGRES = bool(DATABASE_URL)

if IS_POSTGRES:
    import psycopg2
    from psycopg2.extras import RealDictCursor


class PSQLCursorWrapper:
    def __init__(self, cursor):
        self.cursor = cursor

    @property
    def lastrowid(self) -> int:
        cursor2 = self.cursor.connection.cursor()
        cursor2.execute("SELECT lastval()")
        val = cursor2.fetchone()[0]
        cursor2.close()
        return val

    def execute(self, sql: str, params=None):
        sql = sql.replace("?", "%s")
        self.cursor.execute(sql, params)
        return self

    def fetchone(self):
        row = self.cursor.fetchone()
        if row is None:
            return None
        return dict(row)

    def fetchall(self):
        rows = self.cursor.fetchall()
        return [dict(r) for r in rows]


class PSQLConnectionWrapper:
    def __init__(self, conn):
        self.conn = conn

    def execute(self, sql: str, params=None):
        sql = sql.replace("?", "%s")
        cursor = self.conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute(sql, params)
        return PSQLCursorWrapper(cursor)

    def executescript(self, sql: str):
        sql = sql.replace("?", "%s")
        cursor = self.conn.cursor()
        cursor.execute(sql)
        return cursor

    def commit(self):
        self.conn.commit()

    def close(self):
        self.conn.close()


@dataclass(frozen=True)
class ShiftTotals:
    seconds: int
    earnings: float


@contextmanager
def db() -> Iterator[Any]:
    if IS_POSTGRES:
        conn = psycopg2.connect(DATABASE_URL)
        try:
            yield PSQLConnectionWrapper(conn)
            conn.commit()
        finally:
            conn.close()
    else:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        try:
            yield conn
            conn.commit()
        finally:
            conn.close()


def init_db() -> None:
    with db() as conn:
        if IS_POSTGRES:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS shifts (
                    id SERIAL PRIMARY KEY,
                    user_id BIGINT NOT NULL,
                    start_at TEXT NOT NULL,
                    end_at TEXT,
                    hourly_rate REAL NOT NULL DEFAULT 31.5,
                    auto_closed INTEGER NOT NULL DEFAULT 0,
                    paid_out_at TEXT,
                    created_at TEXT NOT NULL,
                    main_task TEXT,
                    energy INTEGER,
                    sens INTEGER,
                    frustracja INTEGER,
                    work_mode TEXT
                );
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS user_settings (
                    user_id BIGINT PRIMARY KEY,
                    hourly_rate REAL NOT NULL,
                    updated_at TEXT NOT NULL
                );
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS breaks (
                    id SERIAL PRIMARY KEY,
                    shift_id INTEGER NOT NULL REFERENCES shifts(id) ON DELETE CASCADE,
                    start_at TEXT NOT NULL,
                    end_at TEXT
                );
                """
            )
            conn.execute("CREATE INDEX IF NOT EXISTS idx_shifts_user_start ON shifts(user_id, start_at)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_breaks_shift ON breaks(shift_id)")
            
            # Dodatkowe weryfikowanie kolumn dla PostgreSQL w przypadku, gdy tabela juz istniala
            cursor = conn.execute("SELECT column_name FROM information_schema.columns WHERE table_name = 'shifts'")
            columns = {row["column_name"] for row in cursor.fetchall()}
            if "main_task" not in columns:
                conn.execute("ALTER TABLE shifts ADD COLUMN main_task TEXT")
            if "energy" not in columns:
                conn.execute("ALTER TABLE shifts ADD COLUMN energy INTEGER")
            if "sens" not in columns:
                conn.execute("ALTER TABLE shifts ADD COLUMN sens INTEGER")
            if "frustracja" not in columns:
                conn.execute("ALTER TABLE shifts ADD COLUMN frustracja INTEGER")
            if "work_mode" not in columns:
                conn.execute("ALTER TABLE shifts ADD COLUMN work_mode TEXT")
        else:
            conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS shifts (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    start_at TEXT NOT NULL,
                    end_at TEXT,
                    hourly_rate REAL NOT NULL DEFAULT 31.5,
                    auto_closed INTEGER NOT NULL DEFAULT 0,
                    paid_out_at TEXT,
                    created_at TEXT NOT NULL,
                    main_task TEXT,
                    energy INTEGER,
                    sens INTEGER,
                    frustracja INTEGER,
                    work_mode TEXT
                );

                CREATE TABLE IF NOT EXISTS user_settings (
                    user_id INTEGER PRIMARY KEY,
                    hourly_rate REAL NOT NULL,
                    updated_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS breaks (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    shift_id INTEGER NOT NULL REFERENCES shifts(id) ON DELETE CASCADE,
                    start_at TEXT NOT NULL,
                    end_at TEXT
                );

                CREATE INDEX IF NOT EXISTS idx_shifts_user_start ON shifts(user_id, start_at);
                CREATE INDEX IF NOT EXISTS idx_breaks_shift ON breaks(shift_id);
                """
            )
            columns = {row["name"] for row in conn.execute("PRAGMA table_info(shifts)").fetchall()}
            if "hourly_rate" not in columns:
                conn.execute("ALTER TABLE shifts ADD COLUMN hourly_rate REAL NOT NULL DEFAULT 31.5")
            if "main_task" not in columns:
                conn.execute("ALTER TABLE shifts ADD COLUMN main_task TEXT")
            if "energy" not in columns:
                conn.execute("ALTER TABLE shifts ADD COLUMN energy INTEGER")
            if "sens" not in columns:
                conn.execute("ALTER TABLE shifts ADD COLUMN sens INTEGER")
            if "frustracja" not in columns:
                conn.execute("ALTER TABLE shifts ADD COLUMN frustracja INTEGER")
            if "work_mode" not in columns:
                conn.execute("ALTER TABLE shifts ADD COLUMN work_mode TEXT")
            conn.execute("UPDATE shifts SET hourly_rate = ? WHERE hourly_rate IS NULL", (HOURLY_RATE,))


def now_local() -> datetime:
    return datetime.now(TZ).replace(microsecond=0)


def iso(dt: datetime) -> str:
    return dt.astimezone(TZ).isoformat()


def parse_dt(value: str) -> datetime:
    return datetime.fromisoformat(value).astimezone(TZ)


def fmt_dt(dt: datetime) -> str:
    return dt.strftime("%Y-%m-%d %H:%M")


def fmt_time(dt: datetime) -> str:
    return dt.strftime("%H:%M")


def fmt_duration(seconds: int) -> str:
    seconds = max(0, int(seconds))
    hours, rem = divmod(seconds, 3600)
    minutes = rem // 60
    return f"{hours}h {minutes:02d}min"


def fmt_money(amount: float) -> str:
    return f"{amount:.2f} zl"


def parse_date(value: str) -> date:
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError("Data musi byc w formacie YYYY-MM-DD, np. 2026-05-11.") from exc


def parse_loose_date(value: str, current: datetime) -> date:
    normalized = value.strip().lower().replace("/", ".")
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d.%m"):
        try:
            parsed = datetime.strptime(normalized, fmt).date()
            if fmt == "%d.%m":
                parsed = parsed.replace(year=current.year)
            return parsed
        except ValueError:
            continue
    raise ValueError("Nie rozumiem daty. Wpisz np. dzis, wczoraj, 2026-05-11 albo 11.05.")


def parse_clock(value: str) -> time:
    original = value.strip()
    value = original.replace(".", ":")
    if ":" not in value:
        value = f"{value}:00"
    try:
        return datetime.strptime(value, "%H:%M").time()
    except ValueError as exc:
        raise ValueError(f"Nie rozumiem godziny '{original}'. Wpisz np. 8, 8:30 albo 08.30.") from exc


def combine(day: date, clock: str) -> datetime:
    return datetime.combine(day, parse_clock(clock), TZ)


def parse_friendly_day(value: str, current: datetime) -> date:
    normalized = value.strip().lower().rstrip(",")
    if normalized in {"dzis", "dziś", "dzisiaj", "today"}:
        return current.date()
    if normalized in {"wczoraj", "yesterday"}:
        return current.date() - timedelta(days=1)
    return parse_loose_date(normalized, current)


def parse_break_ranges(text: str, day: date) -> list[tuple[datetime, datetime]]:
    text = text.strip().lower()
    if text in {"", "brak", "nie", "-", "0"}:
        return []

    ranges: list[tuple[datetime, datetime]] = []
    for raw_part in text.replace(";", ",").split(","):
        part = raw_part.strip()
        if not part:
            continue
        if "-" in part:
            start_s, end_s = [piece.strip() for piece in part.split("-", 1)]
        else:
            pieces = part.split()
            if len(pieces) != 2:
                raise ValueError(f"Nie rozumiem przerwy '{part}'. Wpisz np. 12-12:30 albo 12:00 12:30.")
            start_s, end_s = pieces
        start = combine(day, start_s)
        end = combine(day, end_s)
        if end <= start:
            raise ValueError("Koniec przerwy musi byc po jej starcie.")
        ranges.append((start, end))
    return ranges


def looks_like_date_token(value: str) -> bool:
    return bool(re.fullmatch(r"\d{1,2}[./-]\d{1,2}([./-]\d{2,4})?|\d{4}-\d{1,2}-\d{1,2}", value))


def input_error(exc: ValueError, example: str) -> str:
    return f"{exc}\n\nPrzyklad:\n{example}"


def normalize_shift_text(text: str) -> str:
    text = text.strip().lower()
    replacements = {
        "od ": "",
        " do ": " ",
        "godz.": "",
        "godz": "",
        "przerwy:": " przerwa ",
        "przerwa:": " przerwa ",
        "pauza": "przerwa",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return re.sub(r"\s+", " ", text).strip()


def parse_shift_text(text: str, current: datetime) -> tuple[date, str, str, list[tuple[datetime, datetime]]]:
    text = normalize_shift_text(text)
    if not text:
        raise ValueError("Pusty wpis.")

    break_text = ""
    match = re.search(r"\b(przerwa|przerwy|p)\b", text)
    if match:
        break_text = text[match.end() :].strip()
        text = text[: match.start()].strip()

    parts = text.split()
    day = current.date()
    if len(parts) >= 2:
        try:
            day = parse_friendly_day(parts[0], current)
            parts = parts[1:]
        except ValueError:
            if looks_like_date_token(parts[0]):
                raise
            day = current.date()

    if len(parts) == 1 and "-" in parts[0]:
        start_clock, end_clock = [piece.strip() for piece in parts[0].split("-", 1)]
        if not start_clock or not end_clock:
            raise ValueError("Podaj start i koniec pracy, np. 8-16.")
    elif len(parts) >= 2:
        start_clock, end_clock = parts[0], parts[1]
    else:
        raise ValueError("Brakuje godzin startu i konca.")

    parse_clock(start_clock)
    parse_clock(end_clock)
    breaks = parse_break_ranges(break_text, day)
    return day, start_clock, end_clock, breaks


def insert_manual_shift(
    conn: sqlite3.Connection,
    user_id: int,
    day: date,
    start_clock: str,
    end_clock: str,
    break_ranges: list[tuple[datetime, datetime]],
    current: datetime,
) -> sqlite3.Row:
    start = combine(day, start_clock)
    end = combine(day, end_clock)
    if end <= start:
        raise ValueError("Koniec musi byc po starcie.")
    for br_start, br_end in break_ranges:
        if br_start < start or br_end > end:
            raise ValueError("Przerwa musi miescic sie w czasie zmiany.")

    cursor = conn.execute(
        "INSERT INTO shifts (user_id, start_at, end_at, hourly_rate, created_at) VALUES (?, ?, ?, ?, ?)",
        (user_id, iso(start), iso(end), get_user_rate(conn, user_id), iso(current)),
    )
    shift_id = cursor.lastrowid
    for br_start, br_end in break_ranges:
        conn.execute(
            "INSERT INTO breaks (shift_id, start_at, end_at) VALUES (?, ?, ?)",
            (shift_id, iso(br_start), iso(br_end)),
        )
    shift = get_shift(conn, user_id, shift_id)
    if shift is None:
        raise RuntimeError("Nie udalo sie odczytac dodanej zmiany.")
    return shift


def get_open_shift(conn: sqlite3.Connection, user_id: int) -> sqlite3.Row | None:
    return conn.execute(
        "SELECT * FROM shifts WHERE user_id = ? AND end_at IS NULL ORDER BY start_at DESC LIMIT 1",
        (user_id,),
    ).fetchone()


def get_user_rate(conn: sqlite3.Connection, user_id: int) -> float:
    row = conn.execute(
        "SELECT hourly_rate FROM user_settings WHERE user_id = ?",
        (user_id,),
    ).fetchone()
    return float(row["hourly_rate"]) if row else HOURLY_RATE


def set_user_rate(conn: sqlite3.Connection, user_id: int, rate: float, current: datetime) -> None:
    conn.execute(
        """
        INSERT INTO user_settings (user_id, hourly_rate, updated_at)
        VALUES (?, ?, ?)
        ON CONFLICT(user_id) DO UPDATE SET hourly_rate = excluded.hourly_rate, updated_at = excluded.updated_at
        """,
        (user_id, rate, iso(current)),
    )
    conn.execute(
        "UPDATE shifts SET hourly_rate = ? WHERE user_id = ? AND end_at IS NULL",
        (rate, user_id),
    )


def get_open_break(conn: sqlite3.Connection, shift_id: int) -> sqlite3.Row | None:
    return conn.execute(
        "SELECT * FROM breaks WHERE shift_id = ? AND end_at IS NULL ORDER BY start_at DESC LIMIT 1",
        (shift_id,),
    ).fetchone()


def close_stale_breaks(conn: sqlite3.Connection, user_id: int, current: datetime) -> list[int]:
    closed_shift_ids: list[int] = []
    rows = conn.execute(
        """
        SELECT s.id AS shift_id, b.start_at AS break_start
        FROM shifts s
        JOIN breaks b ON b.shift_id = s.id
        WHERE s.user_id = ? AND s.end_at IS NULL AND b.end_at IS NULL
        """,
        (user_id,),
    ).fetchall()
    for row in rows:
        break_start = parse_dt(row["break_start"])
        if break_start.date() < current.date():
            conn.execute(
                "UPDATE shifts SET end_at = ?, auto_closed = 1 WHERE id = ?",
                (row["break_start"], row["shift_id"]),
            )
            conn.execute("DELETE FROM breaks WHERE shift_id = ? AND end_at IS NULL", (row["shift_id"],))
            closed_shift_ids.append(row["shift_id"])
    return closed_shift_ids


def shift_work_seconds(conn: sqlite3.Connection, shift: sqlite3.Row, until: datetime | None = None) -> int:
    start = parse_dt(shift["start_at"])
    end = parse_dt(shift["end_at"]) if shift["end_at"] else until or now_local()
    total = max(0, int((end - start).total_seconds()))
    breaks = conn.execute("SELECT * FROM breaks WHERE shift_id = ?", (shift["id"],)).fetchall()
    for br in breaks:
        br_start = parse_dt(br["start_at"])
        br_end = parse_dt(br["end_at"]) if br["end_at"] else until or now_local()
        total -= max(0, int((br_end - br_start).total_seconds()))
    
    # Zaokraglamy czas pracy do najblizszych 15 minut (900 sekund) tylko dla zakonczonych zmian
    if shift["end_at"] is not None:
        total = int(round(total / 900) * 900)
    return max(0, total)


def shift_rate(shift: sqlite3.Row) -> float:
    return float(shift["hourly_rate"]) if "hourly_rate" in shift.keys() else HOURLY_RATE


def shift_earnings(conn: sqlite3.Connection, shift: sqlite3.Row, until: datetime | None = None) -> float:
    return shift_work_seconds(conn, shift, until) / 3600 * shift_rate(shift)


def totals_for_shifts(conn: sqlite3.Connection, shifts: Iterable[sqlite3.Row]) -> ShiftTotals:
    shifts = list(shifts)
    seconds = sum(shift_work_seconds(conn, shift) for shift in shifts)
    earnings = sum(shift_earnings(conn, shift) for shift in shifts)
    return ShiftTotals(seconds=seconds, earnings=earnings)


def user_totals(conn: sqlite3.Connection, user_id: int) -> tuple[ShiftTotals, ShiftTotals]:
    completed = conn.execute(
        "SELECT * FROM shifts WHERE user_id = ? AND end_at IS NOT NULL",
        (user_id,),
    ).fetchall()
    unpaid = [shift for shift in completed if shift["paid_out_at"] is None]
    return totals_for_shifts(conn, unpaid), totals_for_shifts(conn, completed)


def get_shift(conn: sqlite3.Connection, user_id: int, shift_id: int) -> sqlite3.Row | None:
    return conn.execute(
        "SELECT * FROM shifts WHERE user_id = ? AND id = ?",
        (user_id, shift_id),
    ).fetchone()


def shift_number(conn: sqlite3.Connection, shift: sqlite3.Row) -> int:
    return int(
        conn.execute(
            """
            SELECT COUNT(*) AS number
            FROM shifts
            WHERE user_id = ?
              AND (start_at < ? OR (start_at = ? AND id <= ?))
            """,
            (shift["user_id"], shift["start_at"], shift["start_at"], shift["id"]),
        ).fetchone()["number"]
    )


def get_shift_by_number(conn: sqlite3.Connection, user_id: int, number: int) -> sqlite3.Row | None:
    if number <= 0:
        return None
    return conn.execute(
        "SELECT * FROM shifts WHERE user_id = ? ORDER BY start_at, id LIMIT 1 OFFSET ?",
        (user_id, number - 1),
    ).fetchone()


def day_bounds(day: date) -> tuple[str, str]:
    start = datetime.combine(day, time.min, TZ)
    end = start + timedelta(days=1)
    return iso(start), iso(end)


async def send(update: Update, text: str, reply_markup: InlineKeyboardMarkup | None = None) -> None:
    if update.effective_message:
        await update.effective_message.reply_text(text, reply_markup=reply_markup, parse_mode="Markdown")


async def start_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    current = now_local()
    with db() as conn:
        closed = close_stale_breaks(conn, user_id, current)
        shift = get_open_shift(conn, user_id)
        if shift:
            open_break = get_open_break(conn, shift["id"])
            if open_break:
                break_start = parse_dt(open_break["start_at"])
                if break_start.date() != current.date():
                    await send(update, "Poprzednia przerwa byla z innego dnia, wiec zamknalem tamta zmiane na godzine rozpoczecia przerwy. Wpisz /start jeszcze raz, aby zaczac nowa zmiane.")
                    return
                conn.execute("UPDATE breaks SET end_at = ? WHERE id = ?", (iso(current), open_break["id"]))
                await send(update, f"Wrociles z przerwy o {fmt_time(current)}. Licze dalej.")
                return
            await send(update, "Praca jest juz rozpoczeta. Uzyj /przerwa albo /koniec.")
            return

        conn.execute(
            "INSERT INTO shifts (user_id, start_at, hourly_rate, created_at) VALUES (?, ?, ?, ?)",
            (user_id, iso(current), get_user_rate(conn, user_id), iso(current)),
        )
        prefix = f"Automatycznie zamknalem zalegla zmiane: {closed}.\n" if closed else ""
        await send(update, f"{prefix}Start pracy zapisany: {fmt_time(current)}.")


async def break_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    current = now_local()
    with db() as conn:
        close_stale_breaks(conn, user_id, current)
        shift = get_open_shift(conn, user_id)
        if not shift:
            await send(update, "Nie widze aktywnej zmiany. Wpisz /start, gdy zaczynasz prace.")
            return
        if get_open_break(conn, shift["id"]):
            await send(update, "Przerwa juz trwa. Wpisz /start, gdy wrocisz.")
            return
        conn.execute("INSERT INTO breaks (shift_id, start_at) VALUES (?, ?)", (shift["id"], iso(current)))
        await send(update, f"Przerwa zapisana od {fmt_time(current)}. Po przerwie wpisz /start.")


RATING_KEYBOARD = ReplyKeyboardMarkup(
    [["1", "2", "3", "4", "5"], ["6", "7", "8", "9", "10"], ["Pomiń"]],
    resize_keyboard=True,
    one_time_keyboard=True,
)

WORK_MODE_KEYBOARD = ReplyKeyboardMarkup(
    [["Solo", "Ludzie", "Klient"], ["Nauka", "Debug"], ["Pomiń"]],
    resize_keyboard=True,
    one_time_keyboard=True,
)


async def end_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id
    current = now_local()
    with db() as conn:
        close_stale_breaks(conn, user_id, current)
        shift = get_open_shift(conn, user_id)
        if not shift:
            await send(update, "Nie widze aktywnej zmiany do zakonczenia.")
            return ConversationHandler.END

        open_break = get_open_break(conn, shift["id"])
        end_at = parse_dt(open_break["start_at"]) if open_break else current
        if open_break:
            conn.execute("DELETE FROM breaks WHERE id = ?", (open_break["id"],))
        conn.execute("UPDATE shifts SET end_at = ? WHERE id = ?", (iso(end_at), shift["id"]))

        shift = get_shift(conn, user_id, shift["id"])
        today = totals_for_shifts(conn, [shift])
        unpaid, total = user_totals(conn, user_id)
        await send(
            update,
            "\n".join(
                [
                    f"Koniec pracy zapisany: {fmt_time(end_at)}.",
                    f"Dzisiaj pracowales: {fmt_duration(today.seconds)}.",
                    f"Dzisiaj zarobiles: {fmt_money(today.earnings)}.",
                    f"Od ostatniej wyplaty: {fmt_money(unpaid.earnings)}.",
                    f"Lacznie zarobiles: {fmt_money(total.earnings)}.",
                ]
            ),
        )
        
        # Zapisz ID zmiany w user_data na potrzeby ankiety
        context.user_data["end_shift_id"] = shift["id"]
        
        await send(
            update,
            "Pytania podsumowujace dzien (mozesz pominac przez /pomin lub anulowac przez /anuluj):\n\n"
            "1. Zadanie glowne: Jakie bylo glowne zadanie dzisiaj? (np. TCP/IP, kamera, dokumentacja)"
        )
        return END_MAIN_TASK


async def end_main_task_received(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    shift_id = context.user_data.get("end_shift_id")
    if shift_id:
        with db() as conn:
            conn.execute("UPDATE shifts SET main_task = ? WHERE id = ?", (text, shift_id))
    
    return await end_main_task_skipped(update, context)


async def end_main_task_skipped(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await send(
        update,
        "2. Energia: Jak oceniasz poziom swojej energii? (Wpisz liczbe od 1 do 10)",
        reply_markup=RATING_KEYBOARD
    )
    return END_ENERGY


async def end_energy_received(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    try:
        val = int(text)
        if not (1 <= val <= 10):
            raise ValueError()
    except ValueError:
        await send(
            update,
            "Ocena musi byc liczba od 1 do 10. Wybierz z klawiatury lub wpisz liczbe:",
            reply_markup=RATING_KEYBOARD
        )
        return END_ENERGY
        
    shift_id = context.user_data.get("end_shift_id")
    if shift_id:
        with db() as conn:
            conn.execute("UPDATE shifts SET energy = ? WHERE id = ?", (val, shift_id))
            
    return await end_energy_skipped(update, context)


async def end_energy_skipped(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await send(
        update,
        "3. Sens: Jak oceniasz sens dzisiejszej pracy? (Wpisz liczbe od 1 do 10)",
        reply_markup=RATING_KEYBOARD
    )
    return END_SENS


async def end_sens_received(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    try:
        val = int(text)
        if not (1 <= val <= 10):
            raise ValueError()
    except ValueError:
        await send(
            update,
            "Ocena musi byc liczba od 1 do 10. Wybierz z klawiatury lub wpisz liczbe:",
            reply_markup=RATING_KEYBOARD
        )
        return END_SENS
        
    shift_id = context.user_data.get("end_shift_id")
    if shift_id:
        with db() as conn:
            conn.execute("UPDATE shifts SET sens = ? WHERE id = ?", (val, shift_id))
            
    return await end_sens_skipped(update, context)


async def end_sens_skipped(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await send(
        update,
        "4. Frustracja: Jak oceniasz poziom dzisiejszej frustracji? (Wpisz liczbe od 1 do 10)",
        reply_markup=RATING_KEYBOARD
    )
    return END_FRUSTRATION


async def end_frustration_received(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    try:
        val = int(text)
        if not (1 <= val <= 10):
            raise ValueError()
    except ValueError:
        await send(
            update,
            "Ocena musi byc liczba od 1 do 10. Wybierz z klawiatury lub wpisz liczbe:",
            reply_markup=RATING_KEYBOARD
        )
        return END_FRUSTRATION
        
    shift_id = context.user_data.get("end_shift_id")
    if shift_id:
        with db() as conn:
            conn.execute("UPDATE shifts SET frustracja = ? WHERE id = ?", (val, shift_id))
            
    return await end_frustration_skipped(update, context)


async def end_frustration_skipped(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await send(
        update,
        "5. Tryb pracy: W jakim trybie pracowales?",
        reply_markup=WORK_MODE_KEYBOARD
    )
    return END_WORK_MODE


async def end_work_mode_received(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    valid_modes = {"Solo", "Ludzie", "Klient", "Nauka", "Debug"}
    matched = None
    for mode in valid_modes:
        if mode.lower() == text.lower():
            matched = mode
            break
            
    if not matched:
        await send(
            update,
            "Nieprawidlowy tryb pracy. Wybierz z klawiatury (Solo, Ludzie, Klient, Nauka, Debug) lub wpisz 'Pomin':",
            reply_markup=WORK_MODE_KEYBOARD
        )
        return END_WORK_MODE
        
    shift_id = context.user_data.get("end_shift_id")
    if shift_id:
        with db() as conn:
            conn.execute("UPDATE shifts SET work_mode = ? WHERE id = ?", (matched, shift_id))
            
    return await end_work_mode_skipped(update, context)


async def end_work_mode_skipped(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await send(update, "Dziekuje! Dane podsumowujace zmiane zostaly zapisane.", reply_markup=ReplyKeyboardRemove())
    context.user_data.pop("end_shift_id", None)
    return ConversationHandler.END


async def cancel_end_survey(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await send(update, "Anulowano podsumowanie dnia. Zmiana zostala zapisana bez dodatkowych metryk.", reply_markup=ReplyKeyboardRemove())
    context.user_data.pop("end_shift_id", None)
    return ConversationHandler.END


async def status_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    current = now_local()
    with db() as conn:
        closed = close_stale_breaks(conn, user_id, current)
        shift = get_open_shift(conn, user_id)
        unpaid, total = user_totals(conn, user_id)
        lines = []
        if closed:
            lines.append(f"Automatycznie zamknalem zalegle zmiany: {closed}.")
        if shift:
            open_break = get_open_break(conn, shift["id"])
            worked = shift_work_seconds(conn, shift, current)
            if open_break:
                lines.append(f"Status: przerwa od {fmt_time(parse_dt(open_break['start_at']))}.")
            else:
                lines.append(f"Status: pracujesz od {fmt_time(parse_dt(shift['start_at']))}.")
            lines.append(f"Dzisiaj naliczone: {fmt_duration(worked)} / {fmt_money(shift_earnings(conn, shift, current))}.")
        else:
            lines.append("Status: brak aktywnej zmiany.")
        lines.append(f"Aktualna stawka: {fmt_money(get_user_rate(conn, user_id))}/h.")
        lines.append(f"Od ostatniej wyplaty: {fmt_money(unpaid.earnings)}.")
        lines.append(f"Lacznie: {fmt_money(total.earnings)}.")
        await send(update, "\n".join(lines))


async def payout_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    current = now_local()
    with db() as conn:
        close_stale_breaks(conn, user_id, current)
        unpaid, total = user_totals(conn, user_id)
        conn.execute(
            "UPDATE shifts SET paid_out_at = ? WHERE user_id = ? AND end_at IS NOT NULL AND paid_out_at IS NULL",
            (iso(current), user_id),
        )
        await send(
            update,
            f"Wyplata zapisana. Zeruje okres.\nRozliczona kwota: {fmt_money(unpaid.earnings)}.\nLacznie w historii: {fmt_money(total.earnings)}.",
        )


async def rate_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    current = now_local()
    with db() as conn:
        close_stale_breaks(conn, user_id, current)
        if not context.args:
            await send(update, f"Aktualna stawka: {fmt_money(get_user_rate(conn, user_id))}/h.")
            return
        try:
            rate = float(context.args[0].replace(",", "."))
        except ValueError:
            await send(update, "Podaj stawke jako liczbe, np. /stawka 31,50")
            return
        if rate <= 0:
            await send(update, "Stawka musi byc wieksza od zera.")
            return
        set_user_rate(conn, user_id, rate, current)
        await send(update, f"Zapisalem stawke: {fmt_money(rate)}/h. Dotyczy kolejnych zmian oraz aktywnej zmiany, jesli teraz pracujesz.")


def month_bounds(month: str | None, current: datetime) -> tuple[date, date, str]:
    if month:
        first = datetime.strptime(month, "%Y-%m").date().replace(day=1)
    else:
        first = current.date().replace(day=1)
    next_month = (first.replace(day=28) + timedelta(days=4)).replace(day=1)
    return first, next_month, first.strftime("%Y-%m")


async def report_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    current = now_local()
    try:
        first, next_month, label = month_bounds(context.args[0] if context.args else None, current)
    except ValueError:
        await send(update, "Miesiac wpisz jako YYYY-MM.\n\nPrzyklad:\n/raport 2026-05")
        return

    with db() as conn:
        close_stale_breaks(conn, user_id, current)
        shifts = conn.execute(
            "SELECT * FROM shifts WHERE user_id = ? AND start_at >= ? AND start_at < ? ORDER BY start_at",
            (user_id, iso(datetime.combine(first, time.min, TZ)), iso(datetime.combine(next_month, time.min, TZ))),
        ).fetchall()
        if not shifts:
            await send(update, f"Brak zmian w miesiacu {label}.")
            return

        totals = totals_for_shifts(conn, shifts)
        daily: dict[str, ShiftTotals] = {}
        for shift in shifts:
            key = parse_dt(shift["start_at"]).strftime("%Y-%m-%d")
            prev = daily.get(key, ShiftTotals(0, 0.0))
            daily[key] = ShiftTotals(
                prev.seconds + shift_work_seconds(conn, shift, current),
                prev.earnings + shift_earnings(conn, shift, current),
            )

        lines = [
            f"Raport {label}",
            f"Dni pracy: {len(daily)}",
            f"Czas: {fmt_duration(totals.seconds)}",
            f"Zarobek: {fmt_money(totals.earnings)}",
            "",
            "Dni:",
        ]
        for day, day_totals in list(daily.items())[:25]:
            lines.append(f"{day}: {fmt_duration(day_totals.seconds)} / {fmt_money(day_totals.earnings)}")
        if len(daily) > 25:
            lines.append(f"... i jeszcze {len(daily) - 25} dni. Pelna lista jest w /excel.")
        await send(update, "\n".join(lines))


def build_excel(conn: sqlite3.Connection, user_id: int, path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Godziny"
    headers = [
        "ID",
        "Data",
        "Start",
        "Koniec",
        "Przerwy",
        "Czas pracy",
        "Godziny",
        "Stawka",
        "Zarobek",
        "Rozliczone",
        "Auto koniec",
        "Zadanie glowne",
        "Energia",
        "Sens",
        "Frustracja",
        "Tryb pracy",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    shifts = conn.execute(
        "SELECT * FROM shifts WHERE user_id = ? ORDER BY start_at",
        (user_id,),
    ).fetchall()
    for shift in shifts:
        start = parse_dt(shift["start_at"])
        end = parse_dt(shift["end_at"]) if shift["end_at"] else None
        breaks = conn.execute("SELECT * FROM breaks WHERE shift_id = ? ORDER BY start_at", (shift["id"],)).fetchall()
        break_text = ", ".join(
            f"{fmt_time(parse_dt(br['start_at']))}-{fmt_time(parse_dt(br['end_at'])) if br['end_at'] else 'trwa'}"
            for br in breaks
        )
        seconds = shift_work_seconds(conn, shift)
        hours = round(seconds / 3600, 2)
        rate = shift_rate(shift)
        
        keys = shift.keys()
        main_task = shift["main_task"] if "main_task" in keys else ""
        energy = shift["energy"] if "energy" in keys else ""
        sens = shift["sens"] if "sens" in keys else ""
        frustracja = shift["frustracja"] if "frustracja" in keys else ""
        work_mode = shift["work_mode"] if "work_mode" in keys else ""

        ws.append(
            [
                shift["id"],
                start.strftime("%Y-%m-%d"),
                fmt_time(start),
                fmt_time(end) if end else "",
                break_text,
                fmt_duration(seconds),
                hours,
                rate,
                round(seconds / 3600 * rate, 2),
                "tak" if shift["paid_out_at"] else "nie",
                "tak" if shift["auto_closed"] else "nie",
                main_task if main_task is not None else "",
                energy if energy is not None else "",
                sens if sens is not None else "",
                frustracja if frustracja is not None else "",
                work_mode if work_mode is not None else "",
            ]
        )

    unpaid, total = user_totals(conn, user_id)
    summary = wb.create_sheet("Podsumowanie")
    summary.append(["Metryka", "Wartosc"])
    summary.append(["Aktualna stawka", f"{get_user_rate(conn, user_id):.2f} zl/h"])
    summary.append(["Od ostatniej wyplaty - czas", fmt_duration(unpaid.seconds)])
    summary.append(["Od ostatniej wyplaty - zarobek", round(unpaid.earnings, 2)])
    summary.append(["Lacznie - czas", fmt_duration(total.seconds)])
    summary.append(["Lacznie - zarobek", round(total.earnings, 2)])
    for cell in summary[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for column in summary.columns:
        width = max(len(str(cell.value or "")) for cell in column) + 2
        summary.column_dimensions[column[0].column_letter].width = min(width, 36)

    for column in ws.columns:
        width = max(len(str(cell.value or "")) for cell in column) + 2
        ws.column_dimensions[column[0].column_letter].width = min(width, 36)
    wb.save(path)


def build_accountant_excel(conn: sqlite3.Connection, user_id: int, path: Path, first: date, next_month: date) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ewidencja"
    headers = ["Data", "Start", "Koniec", "Liczba godzin"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    shifts = conn.execute(
        "SELECT * FROM shifts WHERE user_id = ? AND start_at >= ? AND start_at < ? AND end_at IS NOT NULL ORDER BY start_at",
        (user_id, iso(datetime.combine(first, time.min, TZ)), iso(datetime.combine(next_month, time.min, TZ))),
    ).fetchall()

    total_hours = 0.0
    for shift in shifts:
        start = parse_dt(shift["start_at"])
        end = parse_dt(shift["end_at"])
        hours = round(shift_work_seconds(conn, shift) / 3600, 2)
        total_hours += hours
        ws.append([start.strftime("%Y-%m-%d"), fmt_time(start), fmt_time(end), hours])

    ws.append([])
    ws.append(["Razem", "", "", round(total_hours, 2)])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for column in ws.columns:
        width = max(len(str(cell.value or "")) for cell in column) + 2
        ws.column_dimensions[column[0].column_letter].width = min(width, 24)
    wb.save(path)


async def excel_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    current = now_local()
    with db() as conn:
        close_stale_breaks(conn, user_id, current)
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / f"godziny_{current.strftime('%Y%m%d_%H%M')}.xlsx"
            build_excel(conn, user_id, path)
            if update.effective_message:
                await update.effective_message.reply_document(path.open("rb"), filename=path.name, caption="Eksport godzin do Excela.")


async def accountant_excel_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    current = now_local()
    try:
        first, next_month, label = month_bounds(context.args[0] if context.args else None, current)
    except ValueError:
        await send(update, "Miesiac wpisz jako YYYY-MM.\n\nPrzyklad:\n/ksiegowa 2026-05")
        return

    with db() as conn:
        close_stale_breaks(conn, user_id, current)
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / f"ewidencja_ksiegowa_{label}.xlsx"
            build_accountant_excel(conn, user_id, path, first, next_month)
            if update.effective_message:
                await update.effective_message.reply_document(
                    path.open("rb"),
                    filename=path.name,
                    caption=f"Ewidencja godzin dla ksiegowej: {label}.",
                )


async def quick_add_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    current = now_local()
    text = " ".join(context.args)
    if not text:
        keyboard = [
            [
                InlineKeyboardButton("📅 Dziś 8:00 - 16:00", callback_data="add:dzis:8-16"),
                InlineKeyboardButton("📅 Wczoraj 8:00 - 16:00", callback_data="add:wczoraj:8-16"),
            ],
            [
                InlineKeyboardButton("📅 Dziś 9:00 - 17:00", callback_data="add:dzis:9-17"),
                InlineKeyboardButton("📅 Wczoraj 9:00 - 17:00", callback_data="add:wczoraj:9-17"),
            ],
            [
                InlineKeyboardButton("✍️ Ręcznie (wpisz np. /dodaj wczoraj 8-16)", callback_data="add:custom")
            ]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await send(
            update,
            "*Szybkie dodawanie zmiany:*\nWybierz jedną z opcji poniżej lub wpisz ręcznie np. `/dodaj wczoraj 8-16`:",
            reply_markup=reply_markup,
        )
        return

    try:
        day, start_clock, end_clock, breaks = parse_shift_text(text, current)
        with db() as conn:
            close_stale_breaks(conn, user_id, current)
            shift = insert_manual_shift(conn, user_id, day, start_clock, end_clock, breaks, current)
            await send(update, f"Dodano: {format_shift_line(conn, shift)}")
    except ValueError as exc:
        await send(
            update,
            input_error(exc, "/dodaj wczoraj 8-16 przerwa 12-12:30"),
        )


async def fill_start_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = " ".join(context.args)
    if text:
        return await fill_details_from_text(update, context, text)
    await send(
        update,
        "Napisz brakujaca zmiane jednym zdaniem, np.\n"
        "wczoraj 8-16\n"
        "11.05 8:00-16:00 przerwa 12-12:30\n"
        "dzis od 8 do 15:30\n\n"
        "Anulujesz przez /anuluj.",
    )
    return ASK_DETAILS


async def fill_details_from_text(update: Update, context: ContextTypes.DEFAULT_TYPE, text: str) -> int:
    user_id = update.effective_user.id
    current = now_local()
    try:
        day, start_clock, end_clock, breaks = parse_shift_text(text, current)
        with db() as conn:
            close_stale_breaks(conn, user_id, current)
            shift = insert_manual_shift(conn, user_id, day, start_clock, end_clock, breaks, current)
            await send(update, f"Gotowe, dodalem wpis:\n{format_shift_line(conn, shift)}")
    except ValueError as exc:
        await send(
            update,
            input_error(exc, "/uzupelnij wczoraj 8-16 przerwa 12-12:30"),
        )
        return ASK_DETAILS
    return ConversationHandler.END


async def fill_details_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    return await fill_details_from_text(update, context, update.effective_message.text)


async def cancel_fill(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await send(update, "Anulowalem uzupelnianie.")
    return ConversationHandler.END


def format_shift_line(conn: sqlite3.Connection, shift: sqlite3.Row) -> str:
    number = shift_number(conn, shift)
    start = parse_dt(shift["start_at"])
    end = parse_dt(shift["end_at"]) if shift["end_at"] else None
    seconds = shift_work_seconds(conn, shift)
    breaks = conn.execute("SELECT * FROM breaks WHERE shift_id = ? ORDER BY start_at", (shift["id"],)).fetchall()
    br = ""
    if breaks:
        br = " | przerwy: " + ", ".join(
            f"#{b['id']} {fmt_time(parse_dt(b['start_at']))}-{fmt_time(parse_dt(b['end_at'])) if b['end_at'] else 'trwa'}"
            for b in breaks
        )
    
    details = []
    keys = shift.keys()
    if "main_task" in keys and shift["main_task"]:
        details.append(f"zadanie: {shift['main_task']}")
    if "energy" in keys and shift["energy"] is not None:
        details.append(f"energia: {shift['energy']}/10")
    if "sens" in keys and shift["sens"] is not None:
        details.append(f"sens: {shift['sens']}/10")
    if "frustracja" in keys and shift["frustracja"] is not None:
        details.append(f"frustracja: {shift['frustracja']}/10")
    if "work_mode" in keys and shift["work_mode"]:
        details.append(f"tryb: {shift['work_mode']}")
        
    det_str = ""
    if details:
        det_str = " | " + ", ".join(details)
        
    return f"#{number} {fmt_dt(start)} - {fmt_time(end) if end else 'trwa'} | {fmt_duration(seconds)} | {fmt_money(shift_earnings(conn, shift))}{br}{det_str}"


async def delete_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    args = context.args
    current = now_local()
    with db() as conn:
        close_stale_breaks(conn, user_id, current)
        if not args:
            shifts = conn.execute(
                "SELECT * FROM shifts WHERE user_id = ? ORDER BY start_at DESC LIMIT 5",
                (user_id,),
            ).fetchall()
            if not shifts:
                await send(update, "Nie masz jeszcze żadnych wpisów do usunięcia.")
                return
            
            keyboard = []
            lines = ["*Wybierz zmianę, którą chcesz usunąć (kliknij przycisk poniżej):*", ""]
            for shift in shifts:
                number = shift_number(conn, shift)
                lines.append(format_shift_line(conn, shift))
                keyboard.append([InlineKeyboardButton(f"❌ Usuń #{number}", callback_data=f"del:{shift['id']}:{number}")])
            
            reply_markup = InlineKeyboardMarkup(keyboard)
            await send(update, "\n".join(lines), reply_markup=reply_markup)
            return

        try:
            if args[0].lower() in {"ostatni", "ostatnie"}:
                shift = conn.execute(
                    "SELECT * FROM shifts WHERE user_id = ? ORDER BY start_at DESC LIMIT 1",
                    (user_id,),
                ).fetchone()
            else:
                shift = get_shift_by_number(conn, user_id, int(args[0]))
        except ValueError:
            await send(update, "Podaj numer wpisu, np. `/usun 3` lub `/usun ostatni`.")
            return

        if not shift:
            await send(update, "Nie znalazłem takiego wpisu.")
            return

        deleted = format_shift_line(conn, shift)
        conn.execute("DELETE FROM shifts WHERE id = ? AND user_id = ?", (shift["id"], user_id))
        await send(update, f"Usunięto wpis:\n{deleted}")


async def list_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    current = now_local()
    with db() as conn:
        close_stale_breaks(conn, user_id, current)
        shifts = conn.execute(
            "SELECT * FROM shifts WHERE user_id = ? ORDER BY start_at DESC LIMIT 5",
            (user_id,),
        ).fetchall()
        if not shifts:
            await send(update, "Nie ma jeszcze zadnych wpisow.")
            return
        lines = ["Ostatnie wpisy:"]
        lines.extend(format_shift_line(conn, shift) for shift in shifts)
        await send(update, "\n".join(lines))


async def edit_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    args = context.args
    current = now_local()
    try:
        with db() as conn:
            close_stale_breaks(conn, user_id, current)
            if not args:
                await send(
                    update,
                    "\n".join(
                        [
                            "Uzycie /popraw:",
                            "/popraw lista [YYYY-MM-DD]",
                            "/popraw dodaj YYYY-MM-DD HH:MM HH:MM",
                            "Latwiej: /uzupelnij albo /dodaj wczoraj 8 16 przerwa 12:00-12:30",
                            "/popraw start NUMER HH:MM",
                            "/popraw koniec NUMER HH:MM albo brak",
                            "/popraw usun NUMER",
                            "/popraw przerwa NUMER dodaj HH:MM HH:MM",
                            "/popraw przerwa NUMER usun BREAK_ID",
                            "/popraw zadanie NUMER [tekst albo brak]",
                            "/popraw energia NUMER [1-10 albo brak]",
                            "/popraw sens NUMER [1-10 albo brak]",
                            "/popraw frustracja NUMER [1-10 albo brak]",
                            "/popraw tryb NUMER [Solo/Ludzie/Klient/Nauka/Debug albo brak]",
                        ]
                    ),
                )
                return

            action = args[0].lower()
            if action == "lista":
                day = parse_date(args[1]) if len(args) > 1 else current.date()
                start_s, end_s = day_bounds(day)
                shifts = conn.execute(
                    "SELECT * FROM shifts WHERE user_id = ? AND start_at >= ? AND start_at < ? ORDER BY start_at",
                    (user_id, start_s, end_s),
                ).fetchall()
                if not shifts:
                    await send(update, f"Brak zmian dla {day}.")
                    return
                await send(update, "\n".join(format_shift_line(conn, shift) for shift in shifts))
                return

            if action == "dodaj" and len(args) == 4:
                day = parse_date(args[1])
                start = combine(day, args[2])
                end = combine(day, args[3])
                if end <= start:
                    raise ValueError("Koniec musi byc po starcie.")
                conn.execute(
                    "INSERT INTO shifts (user_id, start_at, end_at, hourly_rate, created_at) VALUES (?, ?, ?, ?, ?)",
                    (user_id, iso(start), iso(end), get_user_rate(conn, user_id), iso(current)),
                )
                await send(update, "Dodano zmiane.")
                return

            if action in {"start", "koniec", "usun", "zadanie", "energia", "sens", "frustracja", "tryb"}:
                shift_number_ref = int(args[1])
                shift = get_shift_by_number(conn, user_id, shift_number_ref)
                if not shift:
                    await send(update, "Nie znaleziono takiej zmiany.")
                    return
                shift_id = shift["id"]
                if action == "usun":
                    conn.execute("DELETE FROM shifts WHERE id = ? AND user_id = ?", (shift_id, user_id))
                    await send(update, "Usunieto zmiane.")
                    return
                if action == "start" and len(args) == 3:
                    old = parse_dt(shift["start_at"])
                    new_start = combine(old.date(), args[2])
                    conn.execute("UPDATE shifts SET start_at = ? WHERE id = ?", (iso(new_start), shift_id))
                    await send(update, "Poprawiono start.")
                    return
                if action == "koniec" and len(args) == 3:
                    if args[2].lower() == "brak":
                        conn.execute("UPDATE shifts SET end_at = NULL, auto_closed = 0 WHERE id = ?", (shift_id,))
                    else:
                        day = parse_dt(shift["start_at"]).date()
                        new_end = combine(day, args[2])
                        conn.execute("UPDATE shifts SET end_at = ?, auto_closed = 0 WHERE id = ?", (iso(new_end), shift_id))
                    await send(update, "Poprawiono koniec.")
                    return
                if action == "zadanie" and len(args) >= 3:
                    task_text = " ".join(args[2:])
                    if task_text.lower() == "brak":
                        conn.execute("UPDATE shifts SET main_task = NULL WHERE id = ?", (shift_id,))
                    else:
                        conn.execute("UPDATE shifts SET main_task = ? WHERE id = ?", (task_text, shift_id))
                    await send(update, "Poprawiono zadanie.")
                    return
                if action == "energia" and len(args) == 3:
                    val_str = args[2]
                    if val_str.lower() == "brak":
                        conn.execute("UPDATE shifts SET energy = NULL WHERE id = ?", (shift_id,))
                    else:
                        val = int(val_str)
                        if not (1 <= val <= 10):
                            raise ValueError("Ocena musi byc od 1 do 10.")
                        conn.execute("UPDATE shifts SET energy = ? WHERE id = ?", (val, shift_id))
                    await send(update, "Poprawiono energie.")
                    return
                if action == "sens" and len(args) == 3:
                    val_str = args[2]
                    if val_str.lower() == "brak":
                        conn.execute("UPDATE shifts SET sens = NULL WHERE id = ?", (shift_id,))
                    else:
                        val = int(val_str)
                        if not (1 <= val <= 10):
                            raise ValueError("Ocena musi byc od 1 do 10.")
                        conn.execute("UPDATE shifts SET sens = ? WHERE id = ?", (val, shift_id))
                    await send(update, "Poprawiono sens.")
                    return
                if action == "frustracja" and len(args) == 3:
                    val_str = args[2]
                    if val_str.lower() == "brak":
                        conn.execute("UPDATE shifts SET frustracja = NULL WHERE id = ?", (shift_id,))
                    else:
                        val = int(val_str)
                        if not (1 <= val <= 10):
                            raise ValueError("Ocena musi byc od 1 do 10.")
                        conn.execute("UPDATE shifts SET frustracja = ? WHERE id = ?", (val, shift_id))
                    await send(update, "Poprawiono frustracje.")
                    return
                if action == "tryb" and len(args) == 3:
                    mode_str = args[2]
                    if mode_str.lower() == "brak":
                        conn.execute("UPDATE shifts SET work_mode = NULL WHERE id = ?", (shift_id,))
                    else:
                        valid_modes = {"Solo", "Ludzie", "Klient", "Nauka", "Debug"}
                        matched = None
                        for mode in valid_modes:
                            if mode.lower() == mode_str.lower():
                                matched = mode
                                break
                        if not matched:
                            raise ValueError("Dozwolone tryby: Solo, Ludzie, Klient, Nauka, Debug.")
                        conn.execute("UPDATE shifts SET work_mode = ? WHERE id = ?", (matched, shift_id))
                    await send(update, "Poprawiono tryb pracy.")
                    return

            if action == "przerwa" and len(args) >= 4:
                shift_number_ref = int(args[1])
                shift = get_shift_by_number(conn, user_id, shift_number_ref)
                if not shift:
                    await send(update, "Nie znaleziono takiej zmiany.")
                    return
                shift_id = shift["id"]
                sub = args[2].lower()
                if sub == "dodaj" and len(args) == 5:
                    day = parse_dt(shift["start_at"]).date()
                    br_start = combine(day, args[3])
                    br_end = combine(day, args[4])
                    if br_end <= br_start:
                        raise ValueError("Koniec przerwy musi byc po jej starcie.")
                    conn.execute(
                        "INSERT INTO breaks (shift_id, start_at, end_at) VALUES (?, ?, ?)",
                        (shift_id, iso(br_start), iso(br_end)),
                    )
                    await send(update, "Dodano przerwe.")
                    return
                if sub == "usun" and len(args) == 4:
                    break_id = int(args[3])
                    conn.execute("DELETE FROM breaks WHERE id = ? AND shift_id = ?", (break_id, shift_id))
                    await send(update, "Usunieto przerwe.")
                    return

        await send(update, "Nie rozumiem tej korekty. Wpisz samo /popraw, zeby zobaczyc przyklady.")
    except ValueError as exc:
        await send(update, f"{exc}\n\nWpisz samo /popraw, zeby zobaczyc przyklady.")
    except IndexError:
        await send(update, "Brakuje danych w komendzie.\n\nWpisz samo /popraw, zeby zobaczyc przyklady.")


async def help_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await send(
        update,
        "\n".join(
            [
                "/start - start pracy albo powrot z przerwy",
                "/przerwa - rozpoczecie przerwy",
                "/koniec - koniec pracy i podsumowanie zarobkow",
                "/status - aktualny stan",
                "/uzupelnij - dodanie zapomnianej zmiany jednym zdaniem",
                "/dodaj - szybkie dodanie, np. /dodaj wczoraj 8-16",
                "/lista - pokazuje 5 ostatnich wpisow",
                "/usun - pokazuje ostatnie wpisy do usuniecia",
                "/usun NUMER - usuwa wpis, np. /usun 3",
                "/raport - podsumowanie miesiaca",
                "/stawka - podglad albo zmiana stawki",
                "/wyplata - wyzerowanie okresu od ostatniej wyplaty",
                "/excel - eksport historii do Excela",
                "/ksiegowa - prosty Excel: data, start, koniec, liczba godzin",
                "/popraw - reczna edycja godzin",
            ]
        ),
    )


async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    if context.error:
        logging.error(
            "Nieoczekiwany blad w bocie",
            exc_info=(type(context.error), context.error, context.error.__traceback__),
        )
    if isinstance(update, Update) and update.effective_message:
        await update.effective_message.reply_text(
            "Cos poszlo nie tak po mojej stronie. Zapisalem blad do bot_errors.log. "
            "Sprobuj jeszcze raz albo wpisz /pomoc."
        )


class HealthCheckHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200)
        self.send_header("Content-type", "text/plain")
        self.end_headers()
        self.wfile.write(b"OK")

    def log_message(self, format, *args):
        # Suppress request logging to keep console/logs clean
        return


def start_health_check_server():
    port = int(os.getenv("PORT", "8080"))
    server = HTTPServer(("0.0.0.0", port), HealthCheckHandler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    logging.info(f"Health check server started on port {port}")


async def callback_query_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()
    user_id = update.effective_user.id
    current = now_local()
    data = query.data

    if data.startswith("del:"):
        _, shift_id_str, number = data.split(":")
        shift_id = int(shift_id_str)
        with db() as conn:
            shift = get_shift(conn, user_id, shift_id)
            if not shift:
                await query.edit_message_text("Nie znaleziono lub zmiana została już usunięta.")
                return
            deleted_line = format_shift_line(conn, shift)
            conn.execute("DELETE FROM shifts WHERE id = ? AND user_id = ?", (shift_id, user_id))
            await query.edit_message_text(f"✅ Pomyślnie usunięto zmianę #{number}!\n\n{deleted_line}")

    elif data.startswith("add:"):
        _, day_token, time_range = data.split(":")
        if day_token == "custom":
            await query.edit_message_text("Wpisz komendę ręcznie, np:\n`/dodaj wczoraj 8:30-16:45` lub `/dodaj dzis 8-16`.")
            return
        text = f"{day_token} {time_range}"
        try:
            day, start_clock, end_clock, breaks = parse_shift_text(text, current)
            with db() as conn:
                close_stale_breaks(conn, user_id, current)
                shift = insert_manual_shift(conn, user_id, day, start_clock, end_clock, breaks, current)
                await query.edit_message_text(f"✅ Pomyślnie dodano zmianę!\n\n{format_shift_line(conn, shift)}")
        except ValueError as exc:
            await query.edit_message_text(f"Błąd podczas szybkiego dodawania: {exc}")


def main() -> None:
    if not BOT_TOKEN or BOT_TOKEN == "wklej_tutaj_token_od_BotFather":
        raise SystemExit("Brak TELEGRAM_BOT_TOKEN w pliku .env.")
    init_db()
    
    # Start healthcheck server for Render
    start_health_check_server()

    app = Application.builder().token(BOT_TOKEN).build()
    fill_handler = ConversationHandler(
        entry_points=[CommandHandler("uzupelnij", fill_start_cmd)],
        states={
            ASK_DETAILS: [MessageHandler(filters.TEXT & ~filters.COMMAND, fill_details_message)],
        },
        fallbacks=[CommandHandler("anuluj", cancel_fill)],
    )
    app.add_handler(fill_handler)

    end_handler = ConversationHandler(
        entry_points=[CommandHandler("koniec", end_cmd)],
        states={
            END_MAIN_TASK: [
                CommandHandler("pomin", end_main_task_skipped),
                MessageHandler(filters.Regex("^(Pomiń|pomiń)$"), end_main_task_skipped),
                MessageHandler(filters.TEXT & ~filters.COMMAND, end_main_task_received),
            ],
            END_ENERGY: [
                CommandHandler("pomin", end_energy_skipped),
                MessageHandler(filters.Regex("^(Pomiń|pomiń)$"), end_energy_skipped),
                MessageHandler(filters.TEXT & ~filters.COMMAND, end_energy_received),
            ],
            END_SENS: [
                CommandHandler("pomin", end_sens_skipped),
                MessageHandler(filters.Regex("^(Pomiń|pomiń)$"), end_sens_skipped),
                MessageHandler(filters.TEXT & ~filters.COMMAND, end_sens_received),
            ],
            END_FRUSTRATION: [
                CommandHandler("pomin", end_frustration_skipped),
                MessageHandler(filters.Regex("^(Pomiń|pomiń)$"), end_frustration_skipped),
                MessageHandler(filters.TEXT & ~filters.COMMAND, end_frustration_received),
            ],
            END_WORK_MODE: [
                CommandHandler("pomin", end_work_mode_skipped),
                MessageHandler(filters.Regex("^(Pomiń|pomiń)$"), end_work_mode_skipped),
                MessageHandler(filters.TEXT & ~filters.COMMAND, end_work_mode_received),
            ],
        },
        fallbacks=[CommandHandler("anuluj", cancel_end_survey)],
    )
    app.add_handler(end_handler)

    app.add_handler(CallbackQueryHandler(callback_query_handler))
    app.add_handler(CommandHandler("start", start_cmd))
    app.add_handler(CommandHandler("przerwa", break_cmd))
    app.add_handler(CommandHandler("status", status_cmd))
    app.add_handler(CommandHandler("dodaj", quick_add_cmd))
    app.add_handler(CommandHandler("lista", list_cmd))
    app.add_handler(CommandHandler("usun", delete_cmd))
    app.add_handler(CommandHandler("raport", report_cmd))
    app.add_handler(CommandHandler("stawka", rate_cmd))
    app.add_handler(CommandHandler("wyplata", payout_cmd))
    app.add_handler(CommandHandler("excel", excel_cmd))
    app.add_handler(CommandHandler("ksiegowa", accountant_excel_cmd))
    app.add_handler(CommandHandler("popraw", edit_cmd))
    app.add_handler(CommandHandler("pomoc", help_cmd))
    app.add_error_handler(error_handler)
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
